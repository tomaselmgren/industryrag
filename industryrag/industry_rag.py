import os
import json
from openai import OpenAI
from dotenv import load_dotenv
import numpy as np
import streamlit as st
from numpy.linalg import norm
from typing import Dict, List
from docx import Document
import openpyxl
import pymupdf
import sqlite3
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_openai import OpenAIEmbeddings
from .prompts import en_prompts

load_dotenv()

class DB:
    def __init__(self, db_file_path, embedding_model_name):
        self
        self.db_file_path = db_file_path
        self.embedding_model_name = embedding_model_name

        self.conn = None

    def connect(self):
        self.conn = sqlite3.connect(self.db_file_path)
    
    def close(self):
        self.conn.commit()
        self.conn.close()

    def create_tables(self):
        cursor = self.conn.cursor()
        cursor.execute('''CREATE TABLE IF NOT EXISTS questions_and_answers 
                            (id INTEGER PRIMARY KEY AUTOINCREMENT, 
                       question TEXT, 
                       answer TEXT, 
                       embedding TEXT, 
                       date TEXT, 
                       temperature REAL, 
                       source_file TEXT, 
                       rating REAL);
                       ''')
        
        cursor.execute('''CREATE TABLE IF NOT EXISTS documents
                       (id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_name TEXT, 
                    full_text TEXT);
                       ''')
        cursor.execute('''CREATE TABLE IF NOT EXISTS chunks (
                       id INTEGER PRIMARY KEY AUTOINCREMENT,
                       document_id INTEGER NOT NULL,
                       chunk_text TEXT NOT NULL,
                       embedding TEXT NOT NULL,
                       FOREIGN KEY (document_id) references documents (id) ON DELETE CASCADE)''')
        self.conn.commit()

    def store_file_and_embeddings(self, filename: str, fulltext: str, chunks_and_embeddings: list[(str, str)]):
        cursor = self.conn.cursor()

        cursor.execute("INSERT INTO documents (file_name, full_text) VALUES (?, ?)", (filename, fulltext))
        doc_id = cursor.lastrowid
    
        for chunk in chunks_and_embeddings:

            embedding = json.dumps(chunk[1]) # Unpack later using json.loads(embedding)

            cursor.execute("INSERT INTO chunks (document_id, chunk_text, embedding) VALUES (?, ?, ?)", (doc_id, chunk[0], embedding))
        
        self.conn.commit()

    def get_top_chunks(self, query_embedding: List[float], top_k: int = 5):
        cursor = self.conn.cursor()

        cursor.execute("""SELECT chunks.chunk_text, chunks.embedding, documents.file_name FROM chunks
                       JOIN documents ON chunks.document_id = documents.id
                       """)
        rows = cursor.fetchall()
        

        chunk_scores = []
        for chunk_text, embedding_str, document_name in rows:
            chunk_embedding = json.loads(embedding_str)
            similarity = self._cos_similarity(query_embedding, chunk_embedding)
            chunk_scores.append((chunk_text, similarity, document_name))

        chunk_scores.sort(key=lambda x: x[1], reverse=True)
        top_chunks = [chunk for chunk in chunk_scores[:top_k]]

        return top_chunks

    def get_file_names_in_database(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT file_name FROM documents")
        rows = cursor.fetchall()
        
        return rows

    def store_question_and_answer(self, question: str, answer: str, date: str, temp: float, source: str):
        
        embedding = self._generate_question_embedding(question)

        cursor = self.conn.cursor()
        cursor.execute('INSERT INTO questions_and_answers (question, answer, embedding, date, temperature, source_file, rating)', (question, answer, embedding, date, temp, source, 0))
        self.conn.commit()

    def fetch_similar_questions_and_answers(self, user_query: str):

        # TODO: This is a highly non-scalable approach. Swap this for some other method.

        embedding = self._generate_question_embedding(user_query)

        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM questions_and_answers")
        rows = cursor.fetchall()

        top_questions = [(row[1], row[2], row[3]) for row in rows if self._cos_similarity(embedding, json.loads(row[3])) >= 0.3]

        def embeddingSort(lst):
            return lst[2]

        top_questions.sort(key=embeddingSort)
        return top_questions[:10]  

    def update_rating(self, id: int, increment: bool, decrement: bool):
        
        value = 0
        if increment == True:
            value = 1
        elif decrement == True:
            value = -1
        
        cursor = self.conn.cursor()
        cursor.execute('SELECT rating FROM questions_and_answers WHERE id = ?', id)
        rating = cursor.fetchone()

        rating += value

        cursor.execute('UPDATE questions_and_answers SET rating = ? WHERE id = ?', rating, id)

        self.conn.commit()

    def _cos_similarity(self, a, b):
        return np.dot(a, b) / (norm(a) * norm(b))

    def _generate_question_embedding(self, question) -> List[float]:
        """
        Generates embedding for a question.
        
        Returns:
        - An embedding of the question
        """
        embeddings_model = OpenAIEmbeddings(model=self.embedding_model_name)

        embedding = embeddings_model.embed_query(question)

        return embedding

        
        
class DocumentToText:
    def __init__(self):
        self,
        #self.api_key = os.getenv("OPENAI_API_KEY")
        self.api_key = st.secrets["OPENAI_API_KEY"]
    
    def process_and_store_documents(self, files, db: DB):
        for file in files:
            temp_file_path = os.path.join("temp", file.name)
            os.makedirs(os.path.dirname(temp_file_path), exist_ok=True)
            with open(temp_file_path, "wb") as f:
                f.write(file.getbuffer())

            document_data = self.extract_text_and_metadata(temp_file_path)
            chunks = self.chunk_document(document_data["Full Text"])
            embeddings = self.generate_embeddings(chunks)
            chunks_and_embeddings = list(zip(chunks, embeddings))
            db.store_file_and_embeddings(document_data["Document Name"], document_data["Full Text"], chunks_and_embeddings)
            

    def extract_text_and_metadata(self, file_path: str) -> Dict:
        """
        Extracts text and metadata from a document.

        Parameters:
        - file_path: Path to the document.

        Returns:
        - A dictionary containing the extracted text and metadata.
        """

        # Initialize the result dictionary
        document_data = {
            "Document Name": os.path.basename(file_path),
            "Document Type": os.path.splitext(file_path)[1].lower(),
            "Full Text": "",
            "Metadata": {
                "File Size": os.path.getsize(file_path),
                "File Type": ""
            }
        }

        # Extract text based on file type
        file_extension = document_data["Document Type"]

        if file_extension == ".txt":
            document_data["Full Text"] = self._handle_txt_file(file_path)

        elif file_extension == ".docx":
            document_data["Full Text"] = self._handle_docx_file(file_path)

        elif file_extension in [".xlsx", ".xls"]:
            document_data["Full Text"] = self._handle_xlsx_file(file_path)

        elif file_extension == ".pdf":
            document_data["Full Text"] = self._handle_pdf_file(file_path)
        else:
            print(f"Unsupported file type: {file_extension}")

        return document_data


    def _handle_txt_file(self, file_path: str) -> str:
        """
        Extracts text from a .txt file.
        """
        with open(file_path, "r", encoding="utf-8") as file:
            return file.read().strip()


    def _handle_docx_file(self, file_path: str) -> str:
        """
        Extracts text from a .docx file.
        """
        doc = Document(file_path)
        return " ".join([para.text for para in doc.paragraphs]).strip()


    def _handle_xlsx_file(self, file_path: str) -> str:
        """
        Extracts text from an .xlsx or .xls file.
        """
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        text = ""

        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                for cell in row:
                    if cell:
                        text += f"{cell} "
            text += "\n"  # Add a line break after each sheet

        return text.strip()
    
    def _handle_pdf_file(self, file_path: str) -> str:
        """
        Extracts text from a .pdf file.
        """
        doc = pymupdf.open(file_path)
        text = ""

        for page in doc:
            text += page.get_textpage().extractTEXT()

        return text.strip()
    
    def chunk_document(self, text: str, chunk_size: int = 512, overlap: int = 50) -> list:
        """
        Chunks the document text into smaller sections.

        Parameters:
        - text: Full document text.
        - chunk_size: Maximum size of each chunk.
        - overlap: Number of overlapping characters between chunks.

        Returns:
        - A list of text chunks.
        """
        # Initialize the text splitter
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=overlap,
            separators=["\n\n", "\n", ".", " ", ""]
        )

        # Split the text
        chunks = text_splitter.split_text(text)

        return chunks

    def generate_embeddings(self, chunks: List[str], model_name: str = "text-embedding-ada-002") -> List[List[float]]:
        """
        Generates embeddings for each chunk of text.

        Parameters:
        - chunks: A list of text chunks to be embedded.
        - model_name: The model to use for generating embeddings. Defaults to "text-embedding-ada-002".

        Returns:
        - A list of embeddings, where each embedding corresponds to a text chunk.
        """
        # Initialize the embeddings model
        embeddings_model = OpenAIEmbeddings(model=model_name, openai_api_key=self.api_key)

        # Generate embeddings for each chunk
        embeddings = [embeddings_model.embed_query(chunk) for chunk in chunks]

        return embeddings

class IndustryRAG:
    def __init__(
        self,
        db: DB,
        model_name: str="gpt-3.5-turbo-0125",
        embeddings_model_name: str="text-embedding-ada-002",
        max_generation_tokens: int = 200,
        temperature: int = 0.5,
    ):
        self.db = db
        self.model_name = model_name
        self.embeddings_model_name = embeddings_model_name
        self.max_generation_tokens = max_generation_tokens
        self.temperature = temperature
        #self.api_key = os.getenv("OPENAI_API_KEY")
        self.api_key = st.secrets["OPENAI_API_KEY"]

        print(f'''Initiated a new IndustryRAG model: \n
              Model: {model_name}, \n
              Embeddings Model: {embeddings_model_name}, \n
              Max Tokens: {max_generation_tokens}, \n
              Temperature: {temperature}\n
              ''')

    def _fetch_relevant_chunks(self, query: str, top_k: int = 5) -> List[str]:
        query_embedding = self._generate_embedding(query)
        top_chunks = self.db.get_top_chunks(query_embedding=query_embedding, top_k=top_k)
        
        relevant_chunks = [
            {"chunk": chunk_text, "document": document_name, "similarity": similarity} 
            for chunk_text, similarity, document_name in top_chunks
        ]

        return relevant_chunks

    def _generate_embedding(self, text: str) -> List[float]:
        embeddings_model = OpenAIEmbeddings(model=self.embeddings_model_name, api_key=self.api_key)
        return embeddings_model.embed_query(text)
    

    def _generate_answer(self, query: str):
        relevant_chunks = self._fetch_relevant_chunks(query)

        response = OpenAI.chat.completions.create(
            model=self.model_name,
            messages=[{
                "role": "system", "content": en_prompts["context"].format(context=relevant_chunks)
            }, {
                "role": "user", "content": en_prompts["qa"].format(question=query)
            }],
            max_tokens=self.max_generation_tokens,
            temperature=self.temperature
        )
        return {"answer": response.choices[0].message.content, "relevant_chunks": relevant_chunks}

    def _print_stats(self):
        return

    def answer(self,
               query):
        
        return self._generate_answer(query)
    
    def rewrite(self,
                query, max_new_tokens: int=128) -> str:
        return
    
    def summarize(self,
                  max_new_tokens: int=512) -> str:
        return 